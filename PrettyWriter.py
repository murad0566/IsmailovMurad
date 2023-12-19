import pandas as pd
import numpy as np
from datetime import datetime, timedelta 
import time
import os
import warnings
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.comments import Comment
import shutil
import xlwings as xw
import re
import hashlib
from adlib import LoadMaster as lm
import pickle
'''
'''



hash_f = hashlib.blake2b(digest_size = 10)


def write_user_hist(f):
    try:
        path_hist = r'D:\Обмен\_Скрипты\Резервные копии рабочих папок\Папка Директора по автоматизации/hist/' + str(datetime.now().date()) +'_'+ os.getlogin() + '.pkl'
        try:
            with open(path_hist, 'rb') as fp:
                t = pickle.load(fp)
        except:
            t = {}

        hash_f.update(bytes(str(time.time()), 'utf-8'))
        t.update({hash_f.hexdigest():{'user':os.getlogin(),'time':str(datetime.now()),'modul':'PrettyWriter','func':f}})
        with open(path_hist, 'wb') as fp:
            pickle.dump(t, fp)
    except:
        pass


upBorder = Border(left=Side(border_style='thin',color='000000'),
                     right=Side(border_style='thin',color='000000'),
                     top=Side(border_style='medium',color='000000'),
                     bottom=Side(border_style='thin',color='000000'))
    
bottomBorder = Border(left=Side(border_style='thin',color='000000'),
                 right=Side(border_style='thin',color='000000'),
                 top=Side(border_style='thin',color='000000'),
                 bottom=Side(border_style='medium',color='000000'))

upLeft = Border(left=Side(border_style='medium',color='000000'),
                 right=Side(border_style='thin',color='000000'),
                 top=Side(border_style='medium',color='000000'),
                 bottom=Side(border_style='thin',color='000000'))

upRight = Border(left=Side(border_style='thin',color='000000'),
                 right=Side(border_style='medium',color='000000'),
                 top=Side(border_style='medium',color='000000'),
                 bottom=Side(border_style='thin',color='000000'))

bottomLeft = Border(left=Side(border_style='medium',color='000000'),
                 right=Side(border_style='thin',color='000000'),
                 top=Side(border_style='thin',color='000000'),
                 bottom=Side(border_style='medium',color='000000'))

bottomRight = Border(left=Side(border_style='thin',color='000000'),
                 right=Side(border_style='medium',color='000000'),
                 top=Side(border_style='thin',color='000000'),
                 bottom=Side(border_style='medium',color='000000'))

Right = Border(left=Side(border_style='thin',color='000000'),
                 right=Side(border_style='medium',color='000000'),
                 top=Side(border_style='thin',color='000000'),
                 bottom=Side(border_style='thin',color='000000'))

myBorder = Border(left=Side(border_style='thin',color='000000'),
                     right=Side(border_style='thin',color='000000'),
                     top=Side(border_style='thin',color='000000'),
                     bottom=Side(border_style='thin',color='000000'))

def help():
    write_user_hist('help')
    print('PrettyWriter - create class')

    print('write_df - write a Dataframe into file')
    print('write_row - write a row into file')
    print('write_col - write a column into file')
    print('write_cell - write a single cell into file')
    
    print('insert_col - insert col by col name')
    print('insert_row - insert row by col name')
    
    print('cells_format - set format to selected cells')
    
    print('set_size - set size to columns or rows')
    print('freeze_panels - freeze rows and cols above the cell')
    print('add_filter - add filter to top row')
    print('merge_cells - merege selected cells')
    print('wrap_cells - wrap text into selected cells')

    print('copy_sheet - copy sheet from one xlsx file to another')
    print('hide - hide rows or columns')
       
def merge_cells(file_name, sheet_name, cells_to_merge):
    '''file_name(str), sheet_name(str), cells_to_merge(str, sample: 'A2:B2')'''
    write_user_hist('merge_cells')   
    book = openpyxl.load_workbook(file_name)
    sheet = book[sheet_name]

    start, end = cells_to_merge.split(':')
    x, start_row = coordinate_from_string(start)
    start_col = column_index_from_string(x)
    x, end_row = coordinate_from_string(end)
    end_col = column_index_from_string(x)
    
    sheet.merge_cells(start_row=start_row, start_column = start_col, end_row=end_row, end_column = end_col)
    book.save(file_name)
      
def add_filter(file_name, sheet_name, filter_range = 'full'):
    '''file_name(str), sheet_name(str), filter_range(str, sample: 'A2:max' or :'B3:G9')'''
    write_user_hist('add_filter') 
    book = openpyxl.load_workbook(file_name)
    sheet = book[sheet_name]
    if filter_range == 'full':
        sheet.auto_filter.ref = sheet.dimensions
    else:
        a,b = filter_range.split(':')
        if b == 'max':
            full_range = a + ':' + get_column_letter(sheet.max_column) + str(sheet.max_row)
            sheet.auto_filter.ref = full_range
        else:
            sheet.auto_filter.ref = filter_range
    book.save(file_name)
    
def freeze_panels(file_name, sheet_name, cell_to_freeze):
    '''file_name(str), sheet_name(str), cell_to_freeze(str, example = "B5")'''
    write_user_hist('freeze_panels') 
    book = openpyxl.load_workbook(file_name)
    sheet = book[sheet_name]
    sheet.freeze_panes = sheet[cell_to_freeze]
    book.save(file_name)
    
def set_size(file_name, sheet_name, range_s, size):
    '''file_name(str), sheet_name(str), range_s(str, sample: "A:B" for columns, '4:7' for rows), size(int, sample: 15)'''
    write_user_hist('set_size') 
    book = openpyxl.load_workbook(file_name)
    sheet = book[sheet_name]
    start, end = range_s.split(':')
    try:
        start_row = int(start)
        end_row = int(end)
        for i in range(start_row, end_row+1):
            sheet.row_dimensions[i].height = size
    except:
        start_col, end_col = column_index_from_string(start), column_index_from_string(end)
        for i in range(start_col, end_col+1):
            sheet.column_dimensions[get_column_letter(i)].width = size
    book.save(file_name)

def cells_format(file_name, sheet_name = 'Лист1', cells = 'A1:B2', cell_color = None , number_format = None, borders = None, al_horizontal = None, al_vertical = None, wrap_text = None, \
                    font_name = None,  font_size = None, font_bold = None, font_italic = None, font_color = None,
                        force_write = False ,save_use = False):
    '''file_name(str), sheet_name(str), cells(str, example = "B5:C10"), cell_color(str) - rgb code (sample 'F0000F'), number_format(str) - set cell format(sample: '0.0%' for percent format (1.2%), 'mmm yyyy' for datetime format (Jan 2021), '# ### ##0' for number split (12 345)\n
    borders - boolean, set borders, al_horizontal - str('fill', 'left', 'center', 'right'), al_vertical - str('center', 'top', 'distributed', 'bottom', 'justify'), wrap_text(boolean) - wrap text into cell, \n
    font_name(str) - fonts's name (sample^ 'Calibri'),  font_size(int) - font's size (sample: 20), font_bold(boolean) - set bold style, font_italic(boolean) - set italic style, font_color(str) - rgb code for font color(sample 'F0000F'),\n
    force_write - boolean, closed your file before writing, save_use - boolean, save and closed your file before writing'''
    write_user_hist('cells_format')  
    try:
        sc, ec = cells.split(':')[0], cells.split(':')[1]
    except:
        sc = ec = cells
    start_row, end_row = int(re.findall(r'\d+', sc)[0]), int(re.findall(r'\d+', ec)[0])
    start_col, end_col = column_index_from_string(sc[:-len(str(start_row))]), column_index_from_string(ec[:-len(str(end_row))])
    
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        if len(book.sheetnames) == 1:
            sheet = book[book.sheetnames[0]]
        else:
            sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)
        
        
    for y in range(start_row, end_row + 1):
        for x in range(start_col, end_col + 1):


            if font_name:
                sheet.cell(row=y, column = x).font = Font(name = font_name)
            if font_size:
                sheet.cell(row=y, column = x).font = Font(size = font_size)
            if font_bold:
                sheet.cell(row=y, column = x).font = Font(bold=font_bold)
            if font_italic:
                sheet.cell(row=y, column = x).font = Font(italic = font_italic)
            if font_color:
                sheet.cell(row=y, column = x).font = Font(color = font_color)
            
            #sheet.cell(row=y, column = x).font = Font(name = font_name, size = font_size, bold=font_bold, italic = font_italic, color = font_color)
            
            if number_format:
                sheet.cell(row = y , column = x).number_format = number_format
    
            if borders:
                sheet.cell(row = y , column = x).border = myBorder
                
            if al_horizontal and al_vertical:
                if wrap_text:
                    sheet.cell(row = y, column = x).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical, wrap_text = wrap_text) 
                else:
                    sheet.cell(row = y, column = x).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical) 
            elif al_vertical:
                if wrap_text:
                    sheet.cell(row = y , column = x).alignment = Alignment(vertical = al_vertical, wrap_text = wrap_text) 
                else:
                    sheet.cell(row = y , column = x).alignment = Alignment(vertical = al_vertical) 
            elif al_horizontal:
                if wrap_text:
                    sheet.cell(row=y, column= x).alignment = Alignment(horizontal = al_horizontal, wrap_text = wrap_text) 
                else:
                    sheet.cell(row=y, column= x).alignment = Alignment(horizontal = al_horizontal) 

            if cell_color:
                sheet.cell(row=y, column = x).fill = PatternFill(patternType ='solid', fgColor = (Color(rgb = cell_color)))
            
    if save_use:
        try:
            book.save(file_name)
        except:
            wb2 = xw.Book(file_name)
            wb2.save()
            if len(xw.books) == 1:
                wb2.app.quit()
            else:
                wb2.close()
            book.save(file_name)
    elif force_write:
        try:
            book.save(file_name)
        except:
            wb2 = xw.Book(file_name)
            if len(xw.books) == 1:
                wb2.app.quit()
            else:
                wb2.close()
            book.save(file_name)
    else:
        book.save(file_name)

def hide(file_name, sheet_name = 'Лист1', range_s = 'A:B'):
    '''file_name(str) - path to file, sheet_name(str) - sheet name,  range_s - range of columns or rows to hide(str, sample: "A:B" for columns, '4:7' for rows)'''
    write_user_hist('hide')
    book = openpyxl.load_workbook(file_name)
    start, end = range_s.split(':')
    sheet = book[sheet_name]
    try:
        start = int(start)
        end = int(end)
        sheet.row_dimensions.group(start, end, hidden = True)
    except:
        sheet.column_dimensions.group(start, end, hidden = True)
    book.save(file_name)

def write_cell(file_name, sheet_name = 'Лист1', data_to_write = None, write_cell = 'A1', cell_color = None , number_format = None, borders = False, al_horizontal = 'right', al_vertical = 'bottom', wrap_text = False, \
                    font_name = 'Calibri',  font_size = 11, font_bold = False, font_italic = False, font_color = '000000',
                        force_write = False ,save_use = False):
    '''file_name(str), sheet_name(str), data_to_write(some variable), write_cell(str, example = "B5"), cell_color(str) - rgb code (sample 'F0000F'), number_format(str) - set cell format(sample: '0.0% for percent format, 'mmm yyyy' for datetime format, '# ### ##0' for number with spaces)\n
    borders - boolean, set borders, al_horizontal - str('fill', 'left', 'center', 'right'), al_vertical - str('center', 'top', 'distributed', 'bottom', 'justify'), wrap_text(boolean) - wrap text into cell, \n
    font_name(str) - fonts's name (sample^ 'Calibri'),  font_size(int) - font's size (sample: 20), font_bold(boolean) - set bold style, font_italic(boolean) - set italic style, font_color(str) - rgb code for font color(sample 'F0000F'),\n
    force_write - boolean, closed your file before writing, save_use - boolean, save and closed your file before writing'''
    write_user_hist('write_cell')
    x, start_row = coordinate_from_string(write_cell)
    start_col = column_index_from_string(x)
    
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)
        
    if number_format:
        sheet.cell(row=start_row, column = start_col).number_format = number_format
   
    if data_to_write:
        sheet.cell(row=start_row, column = start_col).value = data_to_write
    
    sheet.cell(row=start_row, column = start_col).font = Font(name = font_name, size = font_size, bold=font_bold, italic = font_italic, color = font_color)

    if borders:
        sheet.cell(row = start_row , column = start_col).border = myBorder
        
    if al_horizontal and al_vertical:
        if wrap_text:
            sheet.cell(row = start_row, column = start_col).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical, wrap_text = wrap_text) 
        else:
            sheet.cell(row = start_row, column = start_col).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical) 
    elif al_vertical:
        if wrap_text:
            sheet.cell(row = start_row , column = start_col).alignment = Alignment(vertical = al_vertical, wrap_text = wrap_text) 
        else:
            sheet.cell(row = start_row , column = start_col).alignment = Alignment(vertical = al_vertical) 
    elif al_horizontal:
        if wrap_text:
            sheet.cell(row=start_row, column= start_col).alignment = Alignment(horizontal = al_horizontal, wrap_text = wrap_text) 
        else:
            sheet.cell(row=start_row, column= start_col).alignment = Alignment(horizontal = al_horizontal) 

    if cell_color:
        sheet.cell(row=start_row, column = start_col).fill = PatternFill(patternType ='solid', fgColor = (Color(rgb = cell_color)))
    
    if save_use:
        try:
            book.save(file_name)
        except:
            wb2 = xw.Book(file_name)
            wb2.save()
            if len(xw.books) == 1:
                wb2.app.quit()
            else:
                wb2.close()
            book.save(file_name)
    elif force_write:
        try:
            book.save(file_name)
        except:
            wb2 = xw.Book(file_name)
            if len(xw.books) == 1:
                wb2.app.quit()
            else:
                wb2.close()
            book.save(file_name)
    else:
        book.save(file_name)

def write_df(file_name, sheet_name, data_to_write, write_cell = 'A1', write_headers = True, write_index = True, borders = False, al_horizontal = 'right', al_vertical = 'bottom', force_write = False ,save_use = False, force_open = False):
    '''file_name(str), sheet_name(str), data_to_write(pd.Dataframe), write_cell(str, example = "B5"), write_headers(bool), write_index(bool) \n borders - boolean, set borders, al_horizontal and al_vertical - str ('fill', 'left', 'center', 'right'), force_write - boolean, closed your file before writing, save_use - boolean, save and closed your file before writing'''
    write_user_hist('write_df')
    x, start_row = coordinate_from_string(write_cell)
    start_col = column_index_from_string(x)
    d = data_to_write
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)
    if write_headers == True:      
        if write_index == True:
            if type(d.index[0]) == tuple:
                start_col+=2
            else:
                start_col+=1
        try:
            for i in range(len(d.columns)):
                sheet.cell(row=start_row, column = start_col+i).value = d.columns[i]
                sheet.cell(row=start_row, column = start_col+i).border = myBorder
                sheet.cell(row=start_row, column = start_col+i).font = Font(bold=True)
                sheet.cell(row=start_row, column = start_col+i).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                sheet.column_dimensions[get_column_letter(start_col+i)].width = 12
            start_row+=1
        except:
            for i in range(len(d.columns)):
                for j in range(len(d.columns[i])):
                    sheet.cell(row=start_row+j, column = start_col+i).value = d.columns[i][j]
                    sheet.cell(row=start_row+j, column = start_col+i).border = myBorder
                    sheet.cell(row=start_row+j, column = start_col+i).font = Font(bold=True)
                    sheet.cell(row=start_row+j, column = start_col+i).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                    sheet.column_dimensions[get_column_letter(start_col+i)].width = 12        
            sc = 0
            for i in range(len(d.columns)-1):
                if (d.columns[i][0] != d.columns[i+1][0]) or (i == len(d.columns)-2):
                    if i == len(d.columns)-2:
                        if d.columns[i][0] != d.columns[i+1][0]:
                            sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                            ec = i
                            sc = ec+1
                            sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                            break
                        else:
                            i+=1
                    ec = i
                    sheet.merge_cells(start_row=start_row, start_column = start_col + sc, end_row=start_row, end_column = start_col + ec)
                    sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                    sc = ec+1

            start_row+=2
    if write_index == True:         
        if write_headers == True:
            if type(d.index[0]) == tuple:
                start_col -= 2
            else:
                start_col-=1
        if type(d.index[0]) != tuple:
            for i in range(len(d.index)):
                sheet.cell(row=start_row+i, column = start_col).value = d.index[i]
                sheet.cell(row=start_row+i, column = start_col).border = myBorder
                sheet.cell(row=start_row+i, column = start_col).font = Font(bold=True)
                sheet.column_dimensions[get_column_letter(start_col)].width = 15
            start_col+=1
        else:
            for i in range(len(d.index)):
                sheet.cell(row=start_row+i, column = start_col).value = d.index[i][0]
                sheet.cell(row=start_row+i, column = start_col).border = myBorder
                sheet.cell(row=start_row+i, column = start_col).font = Font(bold=True)
                sheet.column_dimensions[get_column_letter(start_col)].width = 15
                
                sheet.cell(row=start_row+i, column = start_col+1).value = d.index[i][1]
                sheet.cell(row=start_row+i, column = start_col+1).border = myBorder
                sheet.cell(row=start_row+i, column = start_col+1).font = Font(bold=True)
                sheet.column_dimensions[get_column_letter(start_col+1)].width = 15
            sc = 0    
            for i in range(len(d.index)-1):
                if (d.index[i][0] != d.index[i+1][0]) or (i == len(d.index)-2):
                    if i == len(d.index)-2:
                        i+=1
                    ec = i
                    sheet.merge_cells(start_row=start_row + sc, start_column = start_col, end_row=start_row + ec, end_column = start_col)
                    sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                    sc = ec+1
            start_col+=2
    for y in range(d.shape[0]):
        for x in range(d.shape[1]):
                sheet.cell(row=start_row + y, column= start_col + x).value = d.iloc[y, x]
                if borders:
                    sheet.cell(row=start_row + y, column= start_col + x).border = myBorder
                if al_horizontal and al_vertical:
                    sheet.cell(row=start_row + y, column= start_col + x).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical) 
                elif al_vertical:
                    sheet.cell(row=start_row + y, column= start_col + x).alignment = Alignment(vertical = al_vertical) 
                elif al_horizontal:
                    sheet.cell(row=start_row + y, column= start_col + x).alignment = Alignment(horizontal = al_horizontal) 
    
    if save_use:
        try:
            book.save(file_name)
        except:
            wb2 = xw.Book(file_name)
            wb2.save()
            if len(xw.books) == 1:
                wb2.app.quit()
            else:
                wb2.close()
            book.save(file_name)
    elif force_write:
        try:
            book.save(file_name)
        except:
            wb2 = xw.Book(file_name)
            if len(xw.books) == 1:
                wb2.app.quit()
            else:
                wb2.close()
            book.save(file_name)
    else:
        book.save(file_name)
        
    if force_open:
        xw.Book(file_name)
             
def write_row(file_name, sheet_name, data_to_write, write_cell):
    '''file_name(str), sheet_name(str), write_cell(str, example = "B5")'''
    write_user_hist('write_row')
    x, start_row = coordinate_from_string(write_cell)
    start_col = column_index_from_string(x)
    d = data_to_write
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)
    for y in range(d.shape[0]):
        sheet.cell(row=start_row , column= start_col+ y).value = d[y]   
    book.save(file_name)   

def insert_col(file_name, sheet_name, col):
    '''file_name(str), sheet_name(str), col(str, example = "B")'''
    write_user_hist('insert_col')   
    x = column_index_from_string(col)
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)
    sheet.insert_cols(x)
    book.save(file_name)  
    
def insert_row(file_name, sheet_name, row):
    '''file_name(str), sheet_name(str), row(int, example = 15)'''
    write_user_hist('insert_row') 
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)
    sheet.insert_rows(row)
    book.save(file_name)    
   
def write_col(file_name, sheet_name, data_to_write, write_cell):
    '''file_name(str), sheet_name(str), write_cell(str, example = "B5")'''
    write_user_hist('write_col') 
    x, start_row = coordinate_from_string(write_cell)
    start_col =column_index_from_string(x)
    d = data_to_write
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)
    for y in range(d.shape[0]):
        sheet.cell(row=start_row + y, column= start_col).value = d[y]   
    book.save(file_name)
  
def wrap_cells(file_name, sheet_name, cells): 
    '''file_name(str), sheet_name(str), cells(str, example = "A2:B14")'''
    write_user_hist('wrap_cells') 
    book = openpyxl.load_workbook(file_name)
    sheet = book[sheet_name]
    sc, ec = cells.split(':')[0], cells.split(':')[1]
    start_row, end_row = int(re.findall(r'\d+', sc)[0]), int(re.findall(r'\d+', ec)[0])
    start_col, end_col = column_index_from_string(sc[:-len(str(start_row))]), column_index_from_string(ec[:-len(str(end_row))])
    for y in range(start_row, end_row + 1):
            for x in range(start_col, end_col + 1):
                    sheet.cell(row = y, column = x).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True)
    book.save(file_name)
    
def copy_sheet(from_file, to_file, sheet_name = 'Sheet1', position = 'start'):
    '''from_file and to_file - str, path to both files, sheet_name - str or list, sheets to copy, position - str, position to copy (start, end or before a sheet (enter a sheet name))'''
    write_user_hist('copy_sheet') 
    wb1 = xw.Book(from_file)
    wb2 = xw.Book(to_file)

    if type(sheet_name) == str:
        sheet_name = [sheet_name]
    for sh in sheet_name:
        ws1 = wb1.sheets(sh)
        if position == 'end':
            to = len(wb2.sheets)
            ws1.api.Copy(After = wb2.sheets(to).api)
        elif position == 'start':
            ws1.api.Copy(Before = wb2.sheets(1).api)
        else:
            ws1.api.Copy(Before = wb2.sheets(position).api)

    wb2.save()
    wb2.close()
    if len(xw.books) == 1:
        wb1.app.quit()
    else:
        wb1.close() 
        
def comment(file_name, sheet_name, text, write_cell, width=False, height=False):
    '''file_name(str), sheet_name(str), text(str), write_cell(str, example = "B5"), width(int, example = 300), height(int, example = 50)'''
    write_user_hist('comment') 
    try:
        book = openpyxl.load_workbook(file_name)
    except:
        book = openpyxl.Workbook()
        for sn in book.sheetnames:
            sh = book.get_sheet_by_name(sn)
            book.remove_sheet(sh)
    try:
        sheet = book[sheet_name]
    except:
        sheet = book.create_sheet(sheet_name)

    comment = Comment(text, 'Author')
    if width==False:
        comment.width = 300
    else:
        comment.width = width

    if height==False:
        comment.height = 50
    else:
        comment.height = height

    sheet[write_cell].comment = comment
    book.save(file_name) 

class PrettyWriter(object):

    ''''file_name(str) - path to file, sheet_name(str) - sheet name. Sample: sheet1 = pw.PrtyyWriter(file_name, sheet_name) sheet1.write_df(data, 'B2') '''
    def __init__(self, file_name, sheet_name):
        write_user_hist('PrettyClass') 
        self.file_name = file_name
        self.sheet_name = sheet_name
        
        
    upBorder = Border(left=Side(border_style='thin',color='000000'),
                         right=Side(border_style='thin',color='000000'),
                         top=Side(border_style='medium',color='000000'),
                         bottom=Side(border_style='thin',color='000000'))

    bottomBorder = Border(left=Side(border_style='thin',color='000000'),
                     right=Side(border_style='thin',color='000000'),
                     top=Side(border_style='thin',color='000000'),
                     bottom=Side(border_style='medium',color='000000'))

    upLeft = Border(left=Side(border_style='medium',color='000000'),
                     right=Side(border_style='thin',color='000000'),
                     top=Side(border_style='medium',color='000000'),
                     bottom=Side(border_style='thin',color='000000'))

    upRight = Border(left=Side(border_style='thin',color='000000'),
                     right=Side(border_style='medium',color='000000'),
                     top=Side(border_style='medium',color='000000'),
                     bottom=Side(border_style='thin',color='000000'))

    bottomLeft = Border(left=Side(border_style='medium',color='000000'),
                     right=Side(border_style='thin',color='000000'),
                     top=Side(border_style='thin',color='000000'),
                     bottom=Side(border_style='medium',color='000000'))

    bottomRight = Border(left=Side(border_style='thin',color='000000'),
                     right=Side(border_style='medium',color='000000'),
                     top=Side(border_style='thin',color='000000'),
                     bottom=Side(border_style='medium',color='000000'))

    Right = Border(left=Side(border_style='thin',color='000000'),
                     right=Side(border_style='medium',color='000000'),
                     top=Side(border_style='thin',color='000000'),
                     bottom=Side(border_style='thin',color='000000'))

    myBorder = Border(left=Side(border_style='thin',color='000000'),
                         right=Side(border_style='thin',color='000000'),
                         top=Side(border_style='thin',color='000000'),
                         bottom=Side(border_style='thin',color='000000'))

    def help(self):
        write_user_hist('help') 
        print('write_df - write a Dataframe into file')
        print('write_row - write a row into file')
        print('write_col - write a column into file')
        print('write_cell - write a single cell into file')

        print('insert_col - insert col by col name')
        print('insert_row - insert col by col name')

        print('cells_format - set format to selected cells')

        print('set_size - set size to columns or rows')
        print('freeze_panels - freeze rows and cols above the cell')
        print('add_filter - add filter to top row')
        print('merge_cells - merege selected cells')
        print('wrap_cells - wrap text into selected cells')

        print('copy_sheet - copy sheet from one xlsx file to another')
        print('hide - hide rows or columns')

    def merge_cells(self, cells_to_merge):
        '''cells_to_merge(str, sample: 'A2:B2')'''
        write_user_hist('merge_cells') 
        book = openpyxl.load_workbook(self.file_name)
        sheet = book[self.sheet_name]

        start, end = cells_to_merge.split(':')
        x, start_row = coordinate_from_string(start)
        start_col = column_index_from_string(x)
        x, end_row = coordinate_from_string(end)
        end_col = column_index_from_string(x)

        sheet.merge_cells(start_row=start_row, start_column = start_col, end_row=end_row, end_column = end_col)
        book.save(self.file_name)

    def add_filter(self, filter_range = 'full'):
        '''filter_range(str, sample: 'A2:max' or :'B3:G9')'''
        write_user_hist('add_filter') 
        book = openpyxl.load_workbook(self.file_name)
        sheet = book[self.sheet_name]
        if filter_range == 'full':
            sheet.auto_filter.ref = sheet.dimensions
        else:
            a,b = filter_range.split(':')
            if b == 'max':
                full_range = a + ':' + get_column_letter(sheet.max_column) + str(sheet.max_row)
                sheet.auto_filter.ref = full_range
            else:
                sheet.auto_filter.ref = filter_range
        book.save(self.file_name)

    def freeze_panels(self, cell_to_freeze):
        '''cell_to_freeze(str, example = "B5")'''
        write_user_hist('freeze_panels') 
        book = openpyxl.load_workbook(self.file_name)
        sheet = book[self.sheet_name]
        sheet.freeze_panes = sheet[cell_to_freeze]
        book.save(self.file_name)

    def set_size(self, range_s, size):
        '''range_s(str, sample: "A:B" for columns, '4:7' for rows), size(int, sample: 15)'''
        write_user_hist('set_size') 
        book = openpyxl.load_workbook(self.file_name)
        sheet = book[self.sheet_name]
        start, end = range_s.split(':')
        try:
            start_row = int(start)
            end_row = int(end)
            for i in range(start_row, end_row+1):
                sheet.row_dimensions[i].height = size
        except:
            start_col, end_col = column_index_from_string(start), column_index_from_string(end)
            for i in range(start_col, end_col+1):
                sheet.column_dimensions[get_column_letter(i)].width = size
        book.save(self.file_name)

    def cells_format(self, cells = 'A1:B2', cell_color = None , number_format = None, borders = None, al_horizontal = None, al_vertical = None, wrap_text = None, \
                        font_name = None,  font_size = None, font_bold = None, font_italic = None, font_color = None,
                            force_write = False ,save_use = False):
        '''cells(str, example = "B5:C10"), cell_color(str) - rgb code (sample 'F0000F'), number_format(str) - set cell format(sample: '0.0%' for percent format (1.2%), 'mmm yyyy' for datetime format (Jan 2021), '# ### ##0' for number split (12 345)\n
        borders - boolean, set borders, al_horizontal - str('fill', 'left', 'center', 'right'), al_vertical - str('center', 'top', 'distributed', 'bottom', 'justify'), wrap_text(boolean) - wrap text into cell, \n
        font_name(str) - fonts's name (sample^ 'Calibri'),  font_size(int) - font's size (sample: 20), font_bold(boolean) - set bold style, font_italic(boolean) - set italic style, font_color(str) - rgb code for font color(sample 'F0000F'),\n
        force_write - boolean, closed your file before writing, save_use - boolean, save and closed your file before writing'''
        write_user_hist('cells_format') 
        try:
            sc, ec = cells.split(':')[0], cells.split(':')[1]
        except:
            sc = ec = cells
        start_row, end_row = int(re.findall(r'\d+', sc)[0]), int(re.findall(r'\d+', ec)[0])
        start_col, end_col = column_index_from_string(sc[:-len(str(start_row))]), column_index_from_string(ec[:-len(str(end_row))])

        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            if len(book.sheetnames) == 1:
                sheet = book[book.sheetnames[0]]
            else:
                sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)


        for y in range(start_row, end_row + 1):
            for x in range(start_col, end_col + 1):

                if font_name:
                    sheet.cell(row=y, column = x).font = Font(name = font_name)
                if font_size:
                    sheet.cell(row=y, column = x).font = Font(size = font_size)
                if font_bold:
                    sheet.cell(row=y, column = x).font = Font(bold=font_bold)
                if font_italic:
                    sheet.cell(row=y, column = x).font = Font(italic = font_italic)
                if font_color:
                    sheet.cell(row=y, column = x).font = Font(color = font_color)

                #sheet.cell(row=y, column = x).font = Font(name = font_name, size = font_size, bold=font_bold, italic = font_italic, color = font_color)

                if number_format:
                    sheet.cell(row = y , column = x).number_format = number_format

                if borders:
                    sheet.cell(row = y , column = x).border = myBorder

                if al_horizontal and al_vertical:
                    if wrap_text:
                        sheet.cell(row = y, column = x).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical, wrap_text = wrap_text) 
                    else:
                        sheet.cell(row = y, column = x).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical) 
                elif al_vertical:
                    if wrap_text:
                        sheet.cell(row = y , column = x).alignment = Alignment(vertical = al_vertical, wrap_text = wrap_text) 
                    else:
                        sheet.cell(row = y , column = x).alignment = Alignment(vertical = al_vertical) 
                elif al_horizontal:
                    if wrap_text:
                        sheet.cell(row=y, column= x).alignment = Alignment(horizontal = al_horizontal, wrap_text = wrap_text) 
                    else:
                        sheet.cell(row=y, column= x).alignment = Alignment(horizontal = al_horizontal) 

                if cell_color:
                    sheet.cell(row=y, column = x).fill = PatternFill(patternType ='solid', fgColor = (Color(rgb = cell_color)))

        if save_use:
            try:
                book.save(self.file_name)
            except:
                wb2 = xw.Book(self.file_name)
                wb2.save()
                if len(xw.books) == 1:
                    wb2.app.quit()
                else:
                    wb2.close()
                book.save(self.file_name)
        elif force_write:
            try:
                book.save(self.file_name)
            except:
                wb2 = xw.Book(self.file_name)
                if len(xw.books) == 1:
                    wb2.app.quit()
                else:
                    wb2.close()
                book.save(self.file_name)
        else:
            book.save(self.file_name)

    def hide(self, range_s = 'A:B'):
        '''range_s - range of columns or rows to hide(str, sample: "A:B" for columns, '4:7' for rows)'''
        write_user_hist('hide') 
        book = openpyxl.load_workbook(self.file_name)
        start, end = range_s.split(':')
        sheet = book[self.sheet_name]
        try:
            start = int(start)
            end = int(end)
            sheet.row_dimensions.group(start, end, hidden = True)
        except:
            sheet.column_dimensions.group(start, end, hidden = True)
        book.save(self.file_name)

    def write_cell(self, data_to_write = None, write_cell = 'A1', cell_color = None , number_format = None, borders = False, al_horizontal = 'right', al_vertical = 'bottom', wrap_text = False, \
                        font_name = 'Calibri',  font_size = 11, font_bold = False, font_italic = False, font_color = '000000',
                            force_write = False ,save_use = False):
        '''data_to_write(some variable), write_cell(str, example = "B5"), cell_color(str) - rgb code (sample 'F0000F'), number_format(str) - set cell format(sample: '0.0% for percent format, 'mmm yyyy' for datetime format, '# ### ##0' for number with spaces)\n
        borders - boolean, set borders, al_horizontal - str('fill', 'left', 'center', 'right'), al_vertical - str('center', 'top', 'distributed', 'bottom', 'justify'), wrap_text(boolean) - wrap text into cell, \n
        font_name(str) - fonts's name (sample^ 'Calibri'),  font_size(int) - font's size (sample: 20), font_bold(boolean) - set bold style, font_italic(boolean) - set italic style, font_color(str) - rgb code for font color(sample 'F0000F'),\n
        force_write - boolean, closed your file before writing, save_use - boolean, save and closed your file before writing'''
        write_user_hist('write_cell') 
        x, start_row = coordinate_from_string(write_cell)
        start_col = column_index_from_string(x)

        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)

        if number_format:
            sheet.cell(row=start_row, column = start_col).number_format = number_format

        if data_to_write:
            sheet.cell(row=start_row, column = start_col).value = data_to_write

        sheet.cell(row=start_row, column = start_col).font = Font(name = font_name, size = font_size, bold=font_bold, italic = font_italic, color = font_color)

        if borders:
            sheet.cell(row = start_row , column = start_col).border = myBorder

        if al_horizontal and al_vertical:
            if wrap_text:
                sheet.cell(row = start_row, column = start_col).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical, wrap_text = wrap_text) 
            else:
                sheet.cell(row = start_row, column = start_col).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical) 
        elif al_vertical:
            if wrap_text:
                sheet.cell(row = start_row , column = start_col).alignment = Alignment(vertical = al_vertical, wrap_text = wrap_text) 
            else:
                sheet.cell(row = start_row , column = start_col).alignment = Alignment(vertical = al_vertical) 
        elif al_horizontal:
            if wrap_text:
                sheet.cell(row=start_row, column= start_col).alignment = Alignment(horizontal = al_horizontal, wrap_text = wrap_text) 
            else:
                sheet.cell(row=start_row, column= start_col).alignment = Alignment(horizontal = al_horizontal) 

        if cell_color:
            sheet.cell(row=start_row, column = start_col).fill = PatternFill(patternType ='solid', fgColor = (Color(rgb = cell_color)))

        if save_use:
            try:
                book.save(self.file_name)
            except:
                wb2 = xw.Book(self.file_name)
                wb2.save()
                if len(xw.books) == 1:
                    wb2.app.quit()
                else:
                    wb2.close()
                book.save(self.file_name)
        elif force_write:
            try:
                book.save(self.file_name)
            except:
                wb2 = xw.Book(self.file_name)
                if len(xw.books) == 1:
                    wb2.app.quit()
                else:
                    wb2.close()
                book.save(self.file_name)
        else:
            book.save(self.file_name)
            
    def comment(self, text, write_cell, width=False, height=False):
        '''file_name(str), sheet_name(str), text(str), write_cell(str, example = "B5"), width(int, example = 300), height(int, example = 50)'''
        write_user_hist('comment') 
        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)

        comment = Comment(text, 'Author')
        if width==False:
            comment.width = 300
        else:
            comment.width = width

        if height==False:
            comment.height = 50
        else:
            comment.height = height

        sheet[write_cell].comment = comment
        book.save(self.file_name) 

    def write_df(self, data_to_write, write_cell = 'A1', write_headers = True, write_index = True, borders = False, al_horizontal = 'right', al_vertical = 'bottom', force_write = False ,save_use = False, force_open = False):
        '''data_to_write(pd.Dataframe), write_cell(str, example = "B5"), write_headers(bool), write_index(bool) \n borders - boolean, set borders, al_horizontal and al_vertical - str ('fill', 'left', 'center', 'right'), force_write - boolean, closed your file before writing, save_use - boolean, save and closed your file before writing'''
        write_user_hist('write_df') 
        x, start_row = coordinate_from_string(write_cell)
        start_col = column_index_from_string(x)
        d = data_to_write
        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)
        if write_headers == True:      
            if write_index == True:
                if type(d.index[0]) == tuple:
                    start_col+=2
                else:
                    start_col+=1
            try:
                for i in range(len(d.columns)):
                    sheet.cell(row=start_row, column = start_col+i).value = d.columns[i]
                    sheet.cell(row=start_row, column = start_col+i).border = myBorder
                    sheet.cell(row=start_row, column = start_col+i).font = Font(bold=True)
                    sheet.cell(row=start_row, column = start_col+i).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                    sheet.column_dimensions[get_column_letter(start_col+i)].width = 12
                start_row+=1
            except:
                for i in range(len(d.columns)):
                    for j in range(len(d.columns[i])):
                        sheet.cell(row=start_row+j, column = start_col+i).value = d.columns[i][j]
                        sheet.cell(row=start_row+j, column = start_col+i).border = myBorder
                        sheet.cell(row=start_row+j, column = start_col+i).font = Font(bold=True)
                        sheet.cell(row=start_row+j, column = start_col+i).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                        sheet.column_dimensions[get_column_letter(start_col+i)].width = 12        
                sc = 0
                for i in range(len(d.columns)-1):
                    if (d.columns[i][0] != d.columns[i+1][0]) or (i == len(d.columns)-2):
                        if i == len(d.columns)-2:
                            if d.columns[i][0] != d.columns[i+1][0]:
                                sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                                ec = i
                                sc = ec+1
                                sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                                break
                            else:
                                i+=1
                        ec = i
                        sheet.merge_cells(start_row=start_row, start_column = start_col + sc, end_row=start_row, end_column = start_col + ec)
                        sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                        sc = ec+1

                start_row+=2
        if write_index == True:         
            if write_headers == True:
                if type(d.index[0]) == tuple:
                    start_col -= 2
                else:
                    start_col-=1
            if type(d.index[0]) != tuple:
                for i in range(len(d.index)):
                    sheet.cell(row=start_row+i, column = start_col).value = d.index[i]
                    sheet.cell(row=start_row+i, column = start_col).border = myBorder
                    sheet.cell(row=start_row+i, column = start_col).font = Font(bold=True)
                    sheet.column_dimensions[get_column_letter(start_col)].width = 15
                start_col+=1
            else:
                for i in range(len(d.index)):
                    sheet.cell(row=start_row+i, column = start_col).value = d.index[i][0]
                    sheet.cell(row=start_row+i, column = start_col).border = myBorder
                    sheet.cell(row=start_row+i, column = start_col).font = Font(bold=True)
                    sheet.column_dimensions[get_column_letter(start_col)].width = 15

                    sheet.cell(row=start_row+i, column = start_col+1).value = d.index[i][1]
                    sheet.cell(row=start_row+i, column = start_col+1).border = myBorder
                    sheet.cell(row=start_row+i, column = start_col+1).font = Font(bold=True)
                    sheet.column_dimensions[get_column_letter(start_col+1)].width = 15
                sc = 0    
                for i in range(len(d.index)-1):
                    if (d.index[i][0] != d.index[i+1][0]) or (i == len(d.index)-2):
                        if i == len(d.index)-2:
                            i+=1
                        ec = i
                        sheet.merge_cells(start_row=start_row + sc, start_column = start_col, end_row=start_row + ec, end_column = start_col)
                        sheet.cell(row=start_row, column = start_col + sc).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True) 
                        sc = ec+1
                start_col+=2
        for y in range(d.shape[0]):
            for x in range(d.shape[1]):
                    sheet.cell(row=start_row + y, column= start_col + x).value = d.iloc[y, x]
                    if borders:
                        sheet.cell(row=start_row + y, column= start_col + x).border = myBorder
                    if al_horizontal and al_vertical:
                        sheet.cell(row=start_row + y, column= start_col + x).alignment = Alignment(horizontal = al_horizontal, vertical = al_vertical) 
                    elif al_vertical:
                        sheet.cell(row=start_row + y, column= start_col + x).alignment = Alignment(vertical = al_vertical) 
                    elif al_horizontal:
                        sheet.cell(row=start_row + y, column= start_col + x).alignment = Alignment(horizontal = al_horizontal) 

        if save_use:
            try:
                book.save(self.file_name)
            except:
                wb2 = xw.Book(self.file_name)
                wb2.save()
                if len(xw.books) == 1:
                    wb2.app.quit()
                else:
                    wb2.close()
                book.save(self.file_name)
        elif force_write:
            try:
                book.save(self.file_name)
            except:
                wb2 = xw.Book(self.file_name)
                if len(xw.books) == 1:
                    wb2.app.quit()
                else:
                    wb2.close()
                book.save(self.file_name)
        else:
            book.save(self.file_name)
            
        if force_open:
            xw.Book(self.file_name)
            
    def write_row(self, data_to_write, write_cell):
        '''data_to_write(DataFrame or Series); write_cell(str, example = "B5")'''
        write_user_hist('write_row') 
        x, start_row = coordinate_from_string(write_cell)
        start_col = column_index_from_string(x)
        d = data_to_write
        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)
        for y in range(d.shape[0]):
            sheet.cell(row=start_row , column= start_col+ y).value = d[y]   
        book.save(self.file_name)   

    def insert_col(self, col):
        '''col(str, example = "B")'''
        write_user_hist('insert_col') 
        x = column_index_from_string(col)
        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)
        sheet.insert_cols(x)
        book.save(self.file_name)  

    def insert_row(self, row):
        '''row(int, example = 15)'''
        write_user_hist('insert_row') 
        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)
        sheet.insert_rows(row)
        book.save(self.file_name)    

    def write_col(self, data_to_write, write_cell):
        '''data_to_write(DataFrame or Series); write_cell(str, example = "B5")'''
        write_user_hist('write_col') 
        x, start_row = coordinate_from_string(write_cell)
        start_col =column_index_from_string(x)
        d = data_to_write
        try:
            book = openpyxl.load_workbook(self.file_name)
        except:
            book = openpyxl.Workbook()
            for sn in book.sheetnames:
                sh = book.get_sheet_by_name(sn)
                book.remove_sheet(sh)
        try:
            sheet = book[self.sheet_name]
        except:
            sheet = book.create_sheet(self.sheet_name)
        for y in range(d.shape[0]):
            sheet.cell(row=start_row + y, column= start_col).value = d[y]   
        book.save(self.file_name)

    def wrap_cells(self, cells): 
        '''cells(str, example = "A2:B14")'''
        write_user_hist('wrap_cells') 
        book = openpyxl.load_workbook(self.file_name)
        sheet = book[self.sheet_name]
        sc, ec = cells.split(':')[0], cells.split(':')[1]
        start_row, end_row = int(re.findall(r'\d+', sc)[0]), int(re.findall(r'\d+', ec)[0])
        start_col, end_col = column_index_from_string(sc[:-len(str(start_row))]), column_index_from_string(ec[:-len(str(end_row))])
        for y in range(start_row, end_row + 1):
                for x in range(start_col, end_col + 1):
                        sheet.cell(row = y, column = x).alignment = Alignment(horizontal='center', vertical ='center', wrap_text = True)
        book.save(self.file_name)
